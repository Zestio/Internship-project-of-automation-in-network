# ============================================================
# NAPALM Tabanlı Multi-Vendor Ağ Yönetim Platformu
# Geliştirici: Zestio | Staj Projesi 2026
# Flask tabanlı web uygulaması — Cisco cihaz yönetimi
# ============================================================

import os
import threading
import time
import hashlib
import subprocess
import platform
import socket
import io
import smtplib
from email.mime.text import MIMEText
from functools import wraps
from flask import Flask, render_template, jsonify, request, redirect, send_file, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime
from database import init_db, log_ekle, log_listele, log_temizle, kullanici_ekle, kullanici_bul, kullanici_cihaz_ekle, kullanici_cihazlari, history_ekle, history_listele
from mock_napalm import get_config, backup_config, compare_config, simulate_config
from compliance import compliance_kontrol
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# ============================================================
# UYGULAMA YAPILANDIRMASI
# ============================================================

app = Flask(__name__)
app.secret_key = "napalm-staj-projesi-2026"  # Session şifreleme anahtarı

# Flask-Login yapılandırması — oturum yönetimi
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"  # Giriş yapılmamışsa yönlendirilecek sayfa

# ============================================================
# KULLANICI MODELİ
# ============================================================

class User(UserMixin):
    """Flask-Login için kullanıcı modeli. Admin ve readonly roller desteklenir."""
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

    @property
    def is_admin(self):
        """Kullanıcı admin mi?"""
        return self.role == 'admin'

    @property
    def is_readonly(self):
        """Kullanıcı sadece görüntüleme yetkisine sahip mi?"""
        return self.role == 'readonly'

@login_manager.user_loader
def load_user(user_id):
    """Flask-Login'in her istekte kullanıcıyı veritabanından yüklemesi için kullanılır."""
    import sqlite3
    conn = sqlite3.connect("audit.db")
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    conn.close()
    if user:
        return User(user[0], user[1], user[3])
    return None

# ============================================================
# DEKORATÖRLER
# ============================================================

def admin_required(f):
    """Sadece admin kullanıcıların erişebileceği route'lar için dekoratör."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('Bu işlem için admin yetkisi gerekli.', 'danger')
            return redirect('/')
        return f(*args, **kwargs)
    return decorated_function

# ============================================================
# VERİTABANI BAŞLATMA
# ============================================================

init_db()
from database import init_device_history
init_device_history()

# ============================================================
# GLOBAL DEĞİŞKENLER
# ============================================================

# Aktif uyarılar listesi ve thread kilidi
alerts = []
alerts_lock = threading.Lock()

# Kullanıcı bazlı cihaz cache'i — {user_id: [cihaz_listesi]}
user_caches = {}
user_caches_lock = threading.Lock()

# E-posta ayarları — /email_settings sayfasından güncellenebilir
EMAIL_SENDER = "berayahya41@gmail.com"
EMAIL_PASSWORD = "wgghdmpjbrpwgszw"
EMAIL_RECEIVER = "naptinsen812@gmail.com"

# ============================================================
# YARDIMCI FONKSİYONLAR
# ============================================================

def sifre_hash(password):
    """Şifreyi SHA-256 ile hashler. Veritabanında düz metin şifre saklanmaz."""
    return hashlib.sha256(password.encode()).hexdigest()

def get_napalm_device(host, port):
    """
    NAPALM ile Cisco IOS cihazına Telnet bağlantısı kurar.
    GNS3 üzerinden Telnet ile bağlandığı için yüksek timeout değerleri gereklidir.
    """
    from napalm import get_network_driver
    driver = get_network_driver('ios')
    return driver(
        hostname=host,
        username="admin",
        password="cisco123",
        optional_args={
            'port': port,
            'transport': 'telnet',
            'secret': "cisco123",
            'global_delay_factor': 30,       # GNS3 gecikmesi için yüksek değer
            'read_timeout_override': 900,    # 15 dakika timeout
            'fast_cli': False,
            'session_timeout': 900,
            'conn_timeout': 120,
        }
    )

def fetch_device(host, port, display_host):
    try:
        # Önce raw socket ile buffer'ı temizle
        import socket as raw_socket
        import time
        s = raw_socket.socket(raw_socket.AF_INET, raw_socket.SOCK_STREAM)
        s.settimeout(10)
        s.connect((host, port))
        time.sleep(3)
        s.settimeout(2)
        try:
            while True:
                data = s.recv(4096)
                if not data:
                    break
        except:
            pass
        s.send(b"\r\n")
        time.sleep(1)
        try:
            s.recv(4096)
        except:
            pass
        s.close()
        time.sleep(2)  # Bağlantının kapanması için bekle

        # Şimdi NAPALM ile bağlan
        dev = get_napalm_device(host, port)
        dev.open()
        facts = dev.get_facts()
        interfaces = dev.get_interfaces()
        try:
            ip_data = dev.get_interfaces_ip()
        except Exception:
            ip_data = {}
        try:
            config = dev.get_config()["running"]
        except Exception:
            config = ""
        dev.close()

        result_interfaces = {}
        for name, info in interfaces.items():
            ip = None
            if name in ip_data:
                ips = list(ip_data[name].get("ipv4", {}).keys())
                ip = ips[0] if ips else None
            result_interfaces[name] = {
                "status": "up" if info["is_up"] else "down",
                "ip": ip
            }

        facts["host"] = display_host
        facts["interfaces"] = result_interfaces
        facts["config"] = config
        return facts

    except Exception as e:
        print(f"HATA ({port}): {e}")
        return {
            "host": display_host,
            "hostname": "Ulaşılamıyor",
            "model": "-",
            "uptime": "-",
            "vendor": "-",
            "os_version": "-",
            "interfaces": {},
            "config": ""
        }

def email_gonder(konu, mesaj):
    """Gmail SMTP üzerinden alarm e-postası gönderir."""
    try:
        msg = MIMEText(mesaj)
        msg['Subject'] = konu
        msg['From'] = EMAIL_SENDER
        msg['To'] = EMAIL_RECEIVER
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
        print(f"Email gonderildi: {konu}")
    except Exception as e:
        print(f"Email HATA: {e}")

def load_user_cache(user_id):
    """
    Kullanıcının tüm cihazlarından paralel olarak veri çeker ve cache'e kaydeder.
    ThreadPoolExecutor ile tüm cihazlara aynı anda bağlanır (paralel).
    """
    devices_db = kullanici_cihazlari(user_id)
    print(f"Cihaz sayisi: {len(devices_db)}")
    if not devices_db:
        print("Kayitli cihaz yok!")
        return

    from concurrent.futures import ThreadPoolExecutor, as_completed
    result = [None] * len(devices_db)

    def fetch_with_index(idx, d):
     try:
        time.sleep(idx * 15)  # 5'ten 15'e çıkar — cihazlar arası daha uzun bekle
        print(f"Cihaz cekiliyor: {d}")
        data = fetch_device(d[2], d[3], d[4])
        return idx, data
     except Exception as e:
        print(f"fetch_with_index HATA (idx={idx}): {e}")
        return idx, {
            "host": d[4],
            "hostname": "Ulaşılamıyor",
            "model": "-",
            "uptime": "-",
            "vendor": "-",
            "os_version": "-",
            "interfaces": {},
            "config": ""
        }

    # Tüm cihazlara paralel bağlan
    with ThreadPoolExecutor(max_workers=1) as executor:
        futures = {executor.submit(fetch_with_index, i, d): i for i, d in enumerate(devices_db)}
        for future in as_completed(futures):
            idx, data = future.result()
            result[idx] = data

    with user_caches_lock:
        user_caches[user_id] = result
    print(f"Cache tamamlandi: {user_id}")

def get_user_devices(user_id):
    """Cache'den kullanıcının cihaz listesini döner. Thread-safe."""
    with user_caches_lock:
        return list(user_caches.get(user_id, []))

def monitor_loop():
    """
    Arka planda çalışan izleme döngüsü.
    Her 5 dakikada bir tüm cihazları kontrol eder.
    Sorun tespit edilirse:
    - Ana sayfada uyarı gösterir
    - E-posta bildirimi gönderir
    - Interface down olaylarını geçmişe kaydeder
    """
    while True:
        time.sleep(300)  # 5 dakika bekle
        try:
            new_alerts = []
            with user_caches_lock:
                all_devices_with_uid = []
                for uid, devices in user_caches.items():
                    for dev in devices:
                        all_devices_with_uid.append((uid, dev))

            for uid, d in all_devices_with_uid:
                # Cihaza ulaşılamıyorsa alarm oluştur
                if d["hostname"] == "Ulaşılamıyor":
                    new_alerts.append({
                        "host": d["host"],
                        "mesaj": f"{d['host']} cihazına ulaşılamıyor!",
                        "tip": "danger"
                    })
                    threading.Thread(
                        target=email_gonder,
                        args=(
                            f"ALARM: {d['host']} cihazina ulasilamiyor",
                            f"{d['host']} adresindeki cihaza ulasilamiyor!\nTarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        ),
                        daemon=True
                    ).start()
                    continue

                # Interface down olan cihazları kontrol et
                for iface, info in d.get("interfaces", {}).items():
                    if info["status"] == "down":
                        new_alerts.append({
                            "host": d["host"],
                            "mesaj": f"{d['hostname']} - {iface} DOWN!",
                            "tip": "warning"
                        })
                        # Geçmişe kaydet
                        history_ekle(uid, d["host"], d["hostname"], iface, "down")
                        # E-posta gönder
                        threading.Thread(
                            target=email_gonder,
                            args=(
                                f"ALARM: {d['hostname']} - {iface} DOWN",
                                f"{d['hostname']} ({d['host']}) cihazinda {iface} DOWN!\nTarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                            ),
                            daemon=True
                        ).start()

            with alerts_lock:
                alerts.clear()
                alerts.extend(new_alerts)

        except Exception as e:
            print(f"Monitor HATA: {e}")

# İzleme döngüsünü arka planda başlat
threading.Thread(target=monitor_loop, daemon=True).start()

def ping_host(host):
    """Belirtilen host'a ping atar. Windows ve Linux için ayrı parametre kullanır."""
    param = "-n" if platform.system().lower() == "windows" else "-c"
    try:
        result = subprocess.run(
            ["ping", param, "1", host],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=3
        )
        return result.returncode == 0
    except Exception:
        return False

def port_scan(host, port, timeout=3):
    """Belirtilen host ve porta TCP bağlantısı denerek port açık mı kontrol eder."""
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((host, port))
        sock.close()
        return result == 0
    except Exception:
        return False

def turkce_temizle(text):
    """PDF export için Türkçe karakterleri ASCII karşılıklarıyla değiştirir.
    ReportLab varsayılan fontları Türkçe karakterleri desteklemez."""
    replacements = {
        'ı': 'i', 'İ': 'I', 'ğ': 'g', 'Ğ': 'G',
        'ü': 'u', 'Ü': 'U', 'ş': 's', 'Ş': 'S',
        'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C',
        '✓': '+', '✗': '-'
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text

# ============================================================
# KİMLİK DOĞRULAMA ROUTE'LARI
# ============================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Kullanıcı giriş sayfası. Başarılı girişte cache arka planda yüklenir."""
    if request.method == 'POST':
        username = request.form.get('username')
        password = sifre_hash(request.form.get('password'))
        user = kullanici_bul(username)
        if user and user[2] == password:
            u = User(user[0], user[1], user[3])
            login_user(u)
            from flask import session
            session['son_giris'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # Cache'i arka planda yükle, kullanıcı hemen ana sayfaya geçsin
            threading.Thread(target=load_user_cache, args=(u.id,), daemon=True).start()
            return redirect('/')
        flash('Kullanıcı adı veya şifre hatalı.', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Yeni kullanıcı kayıt sayfası. Şifre güvenlik kontrolü yapılır."""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        # Kullanıcı adı uzunluk kontrolü
        if len(username) < 3:
            flash('Kullanıcı adı en az 3 karakter olmalı.', 'danger')
            return render_template('register.html')

        # Şifre güvenlik kontrolü: min 8 karakter, büyük harf, rakam
        if len(password) < 8 or not any(c.isupper() for c in password) or not any(c.isdigit() for c in password):
            flash('Şifre en az 8 karakter, 1 büyük harf ve 1 rakam içermeli.', 'danger')
            return render_template('register.html')

        password_hashed = sifre_hash(password)
        host = request.form.get('host')
        port = int(request.form.get('port'))
        display_host = request.form.get('display_host')
        role = request.form.get('role', 'admin')
        user_id = kullanici_ekle(username, password_hashed, role)
        if not user_id:
            flash('Bu kullanıcı adı zaten alınmış.', 'danger')
            return render_template('register.html')
        kullanici_cihaz_ekle(user_id, host, port, display_host)
        flash('Kayıt başarılı, giriş yapabilirsiniz.', 'success')
        return redirect('/login')
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    """Kullanıcıyı oturumdan çıkarır ve giriş sayfasına yönlendirir."""
    logout_user()
    return redirect('/login')

# ============================================================
# ANA SAYFA VE CİHAZ ROUTE'LARI
# ============================================================

@app.route('/')
@login_required
def home():
    """Ana sayfa — kullanıcının cihaz listesini cache'den gösterir."""
    devices = get_user_devices(current_user.id)
    last_scan = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return render_template('index.html', devices=devices, last_scan=last_scan, username=current_user.username)

@app.route('/device/<host>')
@login_required
def device_detail(host):
    """Seçilen cihazın detay sayfası — interface bilgileri ve sistem özeti."""
    devices = get_user_devices(current_user.id)
    device = next((d for d in devices if d["host"] == host), None)
    if not device:
        return "Cihaz bulunamadı", 404
    return render_template('detail.html', facts=device, interfaces=device["interfaces"], host=host)

@app.route('/history/<host>')
@login_required
def device_history(host):
    """Cihazın interface down/up geçmişini gösterir."""
    history = history_listele(current_user.id, host)
    devices = get_user_devices(current_user.id)
    device = next((d for d in devices if d["host"] == host), None)
    hostname = device["hostname"] if device else host
    return render_template('device_history.html', host=host, hostname=hostname, history=history)

# ============================================================
# CONFIG YÖNETİMİ ROUTE'LARI
# ============================================================

@app.route('/config/<host>')
@login_required
def config_page(host):
    """Config yönetimi ana sayfası — mevcut config gösterilir."""
    devices = get_user_devices(current_user.id)
    device = next((d for d in devices if d["host"] == host), None)
    current_config = device.get("config", "") if device else ""
    return render_template('config.html', host=host, current_config=current_config, diff=None, new_config=None)

@app.route('/config/<host>/preview', methods=['POST'])
@login_required
@admin_required
def config_preview(host):
    """Yeni config ile mevcut config arasındaki farkı (diff) gösterir. Sadece admin."""
    new_config = request.form.get('new_config')
    devices = get_user_devices(current_user.id)
    device = next((d for d in devices if d["host"] == host), None)
    current_config = device.get("config", "") if device else ""
    diff = compare_config(host, new_config)
    log_ekle(host, "CONFIG DIFF", "Diff önizlemesi yapıldı", user_id=current_user.id)
    return render_template('config.html', host=host, current_config=current_config, diff=diff, new_config=new_config)

@app.route('/config/<host>/simulate', methods=['POST'])
@login_required
def config_simulate(host):
    """Config simülasyonu — yeni config uygulanırsa interface durumlarını hesaplar.
    Hatalı komutları (geçersiz IP, interface vb.) tespit eder."""
    new_config = request.form.get('new_config')
    interfaces, errors = simulate_config(new_config)
    devices = get_user_devices(current_user.id)
    device = next((d for d in devices if d["host"] == host), None)
    current_config = device.get("config", "") if device else ""
    diff = compare_config(host, new_config)
    return render_template('config.html',
        host=host,
        current_config=current_config,
        diff=diff,
        new_config=new_config,
        sim_interfaces=interfaces,
        sim_errors=errors
    )

@app.route('/config/<host>/apply', methods=['POST'])
@login_required
@admin_required
def config_apply(host):
    """Config değişikliğini uygular ve audit log'a kaydeder. Sadece admin."""
    new_config = request.form.get('new_config')
    log_ekle(host, "CONFIG APPLY", "Config değişikliği uygulandı", user_id=current_user.id)
    return redirect(f'/config/{host}?backup=success')

@app.route('/config/<host>/versions')
@login_required
def config_versions(host):
    """Config versiyonlama sayfası — iki backup arasında diff karşılaştırması yapılabilir."""
    backup_dir = "backups"
    files = []
    if os.path.exists(backup_dir):
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.startswith(host):
                full_path = os.path.join(backup_dir, f)
                size = os.path.getsize(full_path)
                files.append({"name": f, "size": size})
    return render_template('config_versions.html', host=host, files=files)

# ============================================================
# BACKUP ROUTE'LARI
# ============================================================

@app.route('/config/<host>/backup', methods=['POST'])
@login_required
@admin_required
def config_backup(host):
    """Cihazın mevcut config'ini tarih damgalı dosyaya yedekler. Sadece admin."""
    devices = get_user_devices(current_user.id)
    device = next((d for d in devices if d["host"] == host), None)
    config = device.get("config", "") if device else ""
    if not os.path.exists("backups"):
        os.makedirs("backups")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"backups/{host}_{timestamp}.txt"
    with open(filename, "w") as f:
        f.write(config)
    log_ekle(host, "BACKUP", "Config yedeklendi", user_id=current_user.id)
    return redirect(f'/config/{host}?backup=success')

@app.route('/config/<host>/backups')
@login_required
def list_backups(host):
    """Cihaza ait tüm backup dosyalarını listeler."""
    backup_dir = "backups"
    files = []
    if os.path.exists(backup_dir):
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.startswith(host):
                full_path = os.path.join(backup_dir, f)
                size = os.path.getsize(full_path)
                files.append({"name": f, "size": size})
    return render_template('backups.html', host=host, files=files)

@app.route('/config/<host>/backups/download/<filename>')
@login_required
def download_backup(host, filename):
    """Belirtilen backup dosyasını indirir."""
    return send_file(os.path.join("backups", filename), as_attachment=True)

@app.route('/config/<host>/backups/delete/<filename>', methods=['POST'])
@login_required
def delete_backup(host, filename):
    """Belirtilen backup dosyasını kalıcı olarak siler."""
    filepath = os.path.join("backups", filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    return redirect(f'/config/{host}/backups')

# ============================================================
# COMPLIANCE, AUDIT VE İZLEME ROUTE'LARI
# ============================================================

@app.route('/compliance/<host>')
@login_required
def compliance(host):
    """Cihazın güvenlik standartlarına uygunluğunu 10 kural üzerinden denetler."""
    devices = get_user_devices(current_user.id)
    device = next((d for d in devices if d["host"] == host), None)
    config = device.get("config", "") if device else ""
    sonuclar, gecen, toplam = compliance_kontrol(config)
    return render_template('compliance.html', host=host, sonuclar=sonuclar, gecen=gecen, toplam=toplam)

@app.route('/audit')
@login_required
def audit():
    """Kullanıcıya özel audit log sayfası — tüm işlemlerin kaydı."""
    logs = log_listele(current_user.id)
    return render_template('audit.html', logs=logs)

@app.route('/audit/temizle', methods=['POST'])
@login_required
def audit_temizle():
    """Audit log kayıtlarını temizler."""
    log_temizle()
    return redirect('/audit')

@app.route('/topology')
@login_required
def topology():
    """Ağ topoloji haritası — vis.js ile interaktif görselleştirme."""
    devices = get_user_devices(current_user.id)
    # Config gibi büyük alanlar topolojiye gönderilmez
    simple_devices = [{
        "host": d["host"],
        "hostname": d["hostname"],
        "model": d["model"],
        "uptime": d["uptime"],
        "vendor": d.get("vendor", ""),
    } for d in devices]
    return render_template('topology.html', devices=simple_devices)

@app.route('/stats')
@login_required
def stats():
    """İstatistik sayfası — compliance skorları, backup ve audit log grafikleri."""
    devices = get_user_devices(current_user.id)
    logs = log_listele(current_user.id)
    compliance_data = []
    for d in devices:
        if d["hostname"] != "Ulaşılamıyor":
            config = d.get("config", "")
            _, gecen, toplam = compliance_kontrol(config)
            compliance_data.append({
                "hostname": d["hostname"],
                "skor": round((gecen / toplam) * 100),
            })
    backup_count = sum(1 for l in logs if l[3] == "BACKUP")
    diff_count = sum(1 for l in logs if l[3] == "CONFIG DIFF")
    apply_count = sum(1 for l in logs if l[3] == "CONFIG APPLY")
    return render_template('stats.html',
        compliance_data=compliance_data,
        backup_count=backup_count,
        diff_count=diff_count,
        apply_count=apply_count,
        total_devices=len(devices)
    )

# ============================================================
# KULLANICI PROFİL VE HESAP YÖNETİMİ
# ============================================================

@app.route('/profile')
@login_required
def profile():
    """Kullanıcı profil sayfası — rol, son giriş zamanı ve işlem istatistikleri."""
    from flask import session
    devices = get_user_devices(current_user.id)
    logs = log_listele(current_user.id)
    backup_count = sum(1 for l in logs if l[3] == "BACKUP")
    diff_count = sum(1 for l in logs if l[3] == "CONFIG DIFF")
    apply_count = sum(1 for l in logs if l[3] == "CONFIG APPLY")
    son_giris = session.get('son_giris', '-')
    return render_template('profile.html',
        username=current_user.username,
        devices=devices,
        backup_count=backup_count,
        diff_count=diff_count,
        apply_count=apply_count,
        total_devices=len(devices),
        son_giris=son_giris,
        rol=current_user.role
    )

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Kullanıcının şifresini değiştirmesini sağlar. Mevcut şifre doğrulanır."""
    if request.method == 'POST':
        current_password = sifre_hash(request.form.get('current_password'))
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        user = kullanici_bul(current_user.username)
        if user[2] != current_password:
            flash('Mevcut şifre hatalı.', 'danger')
            return render_template('change_password.html')
        if len(new_password) < 8 or not any(c.isupper() for c in new_password) or not any(c.isdigit() for c in new_password):
            flash('Yeni şifre en az 8 karakter, 1 büyük harf ve 1 rakam içermeli.', 'danger')
            return render_template('change_password.html')
        if new_password != confirm_password:
            flash('Yeni şifreler eşleşmiyor.', 'danger')
            return render_template('change_password.html')
        import sqlite3
        conn = sqlite3.connect("audit.db")
        c = conn.cursor()
        c.execute("UPDATE users SET password = ? WHERE id = ?",
                  (sifre_hash(new_password), current_user.id))
        conn.commit()
        conn.close()
        flash('Şifre başarıyla güncellendi.', 'success')
        return redirect('/')
    return render_template('change_password.html')

@app.route('/delete_account', methods=['GET', 'POST'])
@login_required
def delete_account():
    """Kullanıcı hesabını ve tüm verilerini kalıcı olarak siler."""
    if request.method == 'POST':
        password = sifre_hash(request.form.get('password'))
        user = kullanici_bul(current_user.username)
        if user[2] != password:
            flash('Şifre hatalı.', 'danger')
            return render_template('delete_account.html')
        import sqlite3
        conn = sqlite3.connect("audit.db")
        c = conn.cursor()
        # Kullanıcıya ait tüm verileri sil
        c.execute("DELETE FROM user_devices WHERE user_id = ?", (current_user.id,))
        c.execute("DELETE FROM audit_log WHERE user_id = ?", (current_user.id,))
        c.execute("DELETE FROM users WHERE id = ?", (current_user.id,))
        conn.commit()
        conn.close()
        # Cache'den de temizle
        with user_caches_lock:
            user_caches.pop(current_user.id, None)
        logout_user()
        flash('Hesabınız silindi.', 'success')
        return redirect('/login')
    return render_template('delete_account.html')

# ============================================================
# CİHAZ YÖNETİMİ ROUTE'LARI
# ============================================================

@app.route('/add_device', methods=['GET', 'POST'])
@login_required
@admin_required
def add_device():
    """Yeni cihaz ekleme sayfası. Cache arka planda yenilenir. Sadece admin."""
    if request.method == 'POST':
        host = request.form.get('host')
        port = int(request.form.get('port'))
        display_host = request.form.get('display_host')
        kullanici_cihaz_ekle(current_user.id, host, port, display_host)
        threading.Thread(target=load_user_cache, args=(current_user.id,), daemon=True).start()
        flash('Cihaz başarıyla eklendi.', 'success')
        return redirect('/')
    return render_template('add_device.html')

@app.route('/delete_device/<display_host>', methods=['POST'])
@login_required
@admin_required
def delete_device(display_host):
    """Cihazı veritabanından ve cache'den siler. Sadece admin."""
    import sqlite3
    conn = sqlite3.connect("audit.db")
    c = conn.cursor()
    c.execute("DELETE FROM user_devices WHERE user_id = ? AND display_host = ?",
              (current_user.id, display_host))
    conn.commit()
    conn.close()
    # Cache'i de güncelle
    with user_caches_lock:
        if current_user.id in user_caches:
            user_caches[current_user.id] = [
                d for d in user_caches[current_user.id]
                if d["host"] != display_host
            ]
    return redirect('/')

# ============================================================
# E-POSTA VE ADMİN AYARLARI
# ============================================================

@app.route('/email_settings', methods=['GET', 'POST'])
@login_required
def email_settings():
    """E-posta alarm ayarları — gönderici, alıcı ve şifre güncellenir."""
    global EMAIL_SENDER, EMAIL_PASSWORD, EMAIL_RECEIVER
    if request.method == 'POST':
        EMAIL_SENDER = request.form.get('sender')
        EMAIL_PASSWORD = request.form.get('password')
        EMAIL_RECEIVER = request.form.get('receiver')
        flash('Email ayarları güncellendi.', 'success')
        return redirect('/email_settings')
    return render_template('email_settings.html', sender=EMAIL_SENDER, receiver=EMAIL_RECEIVER)

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    """Admin paneli — tüm kullanıcıları ve rollerini listeler."""
    from database import tum_kullanicilar
    users = tum_kullanicilar()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/rol/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def admin_rol_degistir(user_id):
    """Belirtilen kullanıcının rolünü admin veya readonly olarak değiştirir."""
    from database import kullanici_guncelle_rol
    role = request.form.get('role')
    kullanici_guncelle_rol(user_id, role)
    flash('Kullanıcı rolü güncellendi.', 'success')
    return redirect('/admin/users')

# ============================================================
# API ROUTE'LARI
# ============================================================

@app.route('/api/devices')
@login_required
def api_devices():
    """Cache'deki cihaz listesini JSON olarak döner. 'Şimdi Tara' butonu için kullanılır."""
    devices = get_user_devices(current_user.id)
    last_scan = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return jsonify({"devices": devices, "last_scan": last_scan})

@app.route('/api/alerts')
@login_required
def get_alerts():
    """Aktif uyarıları JSON olarak döner. Ana sayfada 30 saniyede bir sorgulanır."""
    with alerts_lock:
        return jsonify(list(alerts))

@app.route('/api/ping/<host>')
@login_required
def ping(host):
    """GNS3 VM IP'sine ping atar. Cihazın erişilebilir olup olmadığını kontrol eder."""
    import sqlite3
    conn = sqlite3.connect("audit.db")
    c = conn.cursor()
    c.execute("SELECT host FROM user_devices WHERE user_id = ? AND display_host = ?",
              (current_user.id, host))
    row = c.fetchone()
    conn.close()
    ping_target = row[0] if row else host
    result = ping_host(ping_target)
    return jsonify({"host": host, "alive": result})

@app.route('/api/portscan/<host>')
@login_required
def portscan(host):
    """Cihazın Telnet portunun açık olup olmadığını kontrol eder."""
    import sqlite3
    conn = sqlite3.connect("audit.db")
    c = conn.cursor()
    c.execute("SELECT host, port FROM user_devices WHERE user_id = ? AND display_host = ?",
              (current_user.id, host))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({"host": host, "port": None, "open": False})
    gns3_host, telnet_port = row[0], row[1]
    result = port_scan(gns3_host, telnet_port)
    return jsonify({"host": host, "port": telnet_port, "open": result})

@app.route('/api/bulk_backup', methods=['POST'])
@login_required
def bulk_backup():
    """Seçili birden fazla cihaza aynı anda backup alır."""
    hosts = request.json.get('hosts', [])
    devices = get_user_devices(current_user.id)
    count = 0
    for host in hosts:
        device = next((d for d in devices if d["host"] == host), None)
        if device:
            config = device.get("config", "")
            if not os.path.exists("backups"):
                os.makedirs("backups")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"backups/{host}_{timestamp}.txt"
            with open(filename, "w") as f:
                f.write(config)
            log_ekle(host, "BACKUP", "Toplu backup alındı", user_id=current_user.id)
            count += 1
    return jsonify({"message": f"{count} cihazdan backup alındı."})

@app.route('/api/bulk_compliance', methods=['POST'])
@login_required
def bulk_compliance():
    """Seçili birden fazla cihaz için compliance denetimi yapar ve sonuçları döner."""
    hosts = request.json.get('hosts', [])
    devices = get_user_devices(current_user.id)
    results = []
    for host in hosts:
        device = next((d for d in devices if d["host"] == host), None)
        if device:
            config = device.get("config", "")
            _, gecen, toplam = compliance_kontrol(config)
            results.append({
                "host": host,
                "hostname": device.get("hostname", "-"),
                "skor": round((gecen / toplam) * 100)
            })
    return jsonify({"results": results})

@app.route('/api/version_diff/<host>', methods=['POST'])
@login_required
def version_diff(host):
    """İki backup dosyası arasındaki farkı (diff) hesaplar ve döner."""
    data = request.json
    v1, v2 = data.get('v1'), data.get('v2')
    try:
        with open(os.path.join("backups", v1), "r") as f:
            config1 = f.read().splitlines()
        with open(os.path.join("backups", v2), "r") as f:
            config2 = f.read().splitlines()
    except Exception as e:
        return jsonify({"diff": f"Hata: {e}"})
    diff_lines = []
    for line in config2:
        if line.strip() not in [l.strip() for l in config1]:
            diff_lines.append(f"+ {line}")
    for line in config1:
        if line.strip() not in [l.strip() for l in config2]:
            diff_lines.append(f"- {line}")
    diff = "\n".join(diff_lines) if diff_lines else "Değişiklik yok"
    return jsonify({"diff": diff})

# ============================================================
# EXPORT ROUTE'LARI
# ============================================================

@app.route('/export/devices/excel')
@login_required
def export_devices_excel():
    """Cihaz listesini Excel (.xlsx) formatında dışa aktarır."""
    from openpyxl.utils import get_column_letter
    devices = get_user_devices(current_user.id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cihaz Listesi"

    # Başlık satırı
    ws.merge_cells("A1:F1")
    ws["A1"] = "Ag Cihaz Envanteri"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Kolon başlıkları
    headers = ["Host", "Hostname", "Vendor", "Model", "OS Version", "Uptime"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2E75B6")
        cell.alignment = Alignment(horizontal="center")

    # Veri satırları
    for row, d in enumerate(devices, start=3):
        ws.cell(row=row, column=1, value=d.get("host", "-"))
        ws.cell(row=row, column=2, value=d.get("hostname", "-"))
        ws.cell(row=row, column=3, value=d.get("vendor", "-"))
        ws.cell(row=row, column=4, value=d.get("model", "-"))
        ws.cell(row=row, column=5, value=d.get("os_version", "-"))
        ws.cell(row=row, column=6, value=str(d.get("uptime", "-")))

    # Kolon genişliklerini otomatik ayarla
    for i in range(1, 7):
        col_letter = get_column_letter(i)
        max_len = max(
            len(str(ws.cell(row=r, column=i).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="cihaz_listesi.xlsx"
    )

@app.route('/export/compliance/pdf')
@login_required
def export_compliance_pdf():
    """Compliance denetim raporunu PDF formatında dışa aktarır."""
    devices = get_user_devices(current_user.id)
    styles = getSampleStyleSheet()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []

    # Rapor başlığı
    elements.append(Paragraph(turkce_temizle("Compliance Denetim Raporu"), styles['Title']))
    elements.append(Paragraph(turkce_temizle(f"Kullanici: {current_user.username}"), styles['Normal']))
    elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Her cihaz için compliance tablosu
    for d in devices:
        if d["hostname"] == "Ulasilamiyor":
            continue
        config = d.get("config", "")
        sonuclar, gecen, toplam = compliance_kontrol(config)
        skor = round((gecen / toplam) * 100)
        elements.append(Paragraph(f"Cihaz: {d['hostname']} ({d['host']})", styles['Heading2']))
        elements.append(Paragraph(f"Compliance Skoru: %{skor}", styles['Normal']))
        elements.append(Spacer(1, 10))
        table_data = [["#", "Kural", "Durum"]]
        for s in sonuclar:
            durum = "+ Uyuyor" if s["durum"] else "- Uymuyor"
            table_data.append([str(s["id"]), turkce_temizle(s["kural"]), durum])
        t = Table(table_data, colWidths=[30, 320, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

    doc.build(elements)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="rapor.pdf"
    )

# ============================================================
# UYGULAMA BAŞLATMA
# ============================================================

if __name__ == '__main__':
    app.run(debug=False)